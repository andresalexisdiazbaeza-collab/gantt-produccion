from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _looks_like_header(a: str, b: str) -> bool:
    al, bl = a.lower(), b.lower()
    return al in {"title", "titulo", "título", "titulo"} and bl in {"material", "materia", "raw material", "materia prima"}


def parse_title_material_excel(path: str | Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows: list[dict[str, str]] = []
    seen: set[tuple[str, str]] = set()
    first = True
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        titulo = _cell_str(row[0] if row else None)
        material = _cell_str(row[1] if row and len(row) > 1 else None)
        if not titulo and not material:
            continue
        if first and _looks_like_header(titulo, material):
            first = False
            continue
        first = False
        if not titulo or not material:
            continue
        key = (titulo, material)
        if key in seen:
            continue
        seen.add(key)
        rows.append({"titulo": titulo, "material": material.upper()})
    wb.close()
    return rows
