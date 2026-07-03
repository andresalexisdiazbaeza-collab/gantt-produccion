from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any, List, Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from ..database import get_db
from ..deps import get_current_user
from ..models import ItemStatus, ProductionItem, User
from ..optimize_service import build_optimization_preview
from ..services import item_to_dict

router = APIRouter(prefix="/export", tags=["export"])


def _wb_response(wb: Workbook, filename: str) -> StreamingResponse:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _write_sheet(ws, headers: List[str], rows: List[List[Any]]) -> None:
    bold = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold
    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            if isinstance(val, (date, datetime)):
                ws.cell(row=r_idx, column=c_idx, value=val.isoformat() if val else "")
            else:
                ws.cell(row=r_idx, column=c_idx, value=val)


ORDER_HEADERS = [
    "Orden", "Cliente", "Material", "Título", "Color", "Matriz mm", "Mallas", "Piezas",
    "Long. pieza", "Kg", "Entregado", "Fecha entrega", "Máquina", "Fecha inicio",
    "Fecha fin", "Días hábiles", "Total m", "M terminados", "M pendientes", "Shrinking", "Estado", "Comentarios", "Notas",
]


def _order_row(item: dict) -> List[Any]:
    return [
        item.get("order_number"),
        item.get("customer"),
        item.get("raw_material"),
        item.get("titulo"),
        item.get("color"),
        item.get("matriz_mm"),
        item.get("meshes"),
        item.get("pieces"),
        item.get("piece_length"),
        item.get("kg_totales"),
        item.get("delivered"),
        item.get("delivery_date"),
        item.get("machine_name"),
        item.get("start_date"),
        item.get("finish_date"),
        item.get("working_days"),
        item.get("total_length"),
        item.get("meters_produced"),
        item.get("remaining_length"),
        item.get("shrinking"),
        item.get("status"),
        item.get("comments"),
        item.get("notes"),
    ]


@router.get("/orders")
def export_orders(
    status: Optional[str] = Query(None, description="activa | terminada"),
    db: Session = Depends(get_db),
    _user: User = Depends(get_current_user),
):
    q = db.query(ProductionItem)
    if status:
        q = q.filter(ProductionItem.status == status)
    items = q.order_by(ProductionItem.order_number).all()
    rows = [_order_row(item_to_dict(i)) for i in items]

    wb = Workbook()
    ws = wb.active
    ws.title = "Ordenes"
    _write_sheet(ws, ORDER_HEADERS, rows)
    suffix = status or "todas"
    return _wb_response(wb, f"ordenes_{suffix}_{date.today().isoformat()}.xlsx")


@router.get("/gantt")
def export_gantt(db: Session = Depends(get_db), _user: User = Depends(get_current_user)):
    items = (
        db.query(ProductionItem)
        .filter(ProductionItem.status == ItemStatus.ACTIVA.value)
        .order_by(ProductionItem.machine_id, ProductionItem.start_date)
        .all()
    )
    rows = [_order_row(item_to_dict(i)) for i in items]

    wb = Workbook()
    ws = wb.active
    ws.title = "Gantt"
    _write_sheet(ws, ORDER_HEADERS, rows)
    return _wb_response(wb, f"gantt_{date.today().isoformat()}.xlsx")


@router.get("/dashboard")
def export_dashboard(db: Session = Depends(get_db), _user: User = Depends(get_current_user)):
    from ..routers.dashboard import get_dashboard_stats

    stats = get_dashboard_stats(db)
    wb = Workbook()

    ws0 = wb.active
    ws0.title = "KPIs"
    _write_sheet(ws0, ["Indicador", "Valor"], [
        ["Órdenes activas", stats.active_count],
        ["Terminadas", stats.completed_count],
        ["Máquinas activas", stats.machines_active],
        ["Kg planificados", stats.total_planned_kg],
        ["Kg producidos", stats.total_produced_kg],
        ["Kg pendientes", stats.total_remaining_kg],
        ["Metros planificados", stats.total_planned_meters],
        ["A tiempo", stats.delivery_compliance["on_time"]],
        ["Tardíos", stats.delivery_compliance["late"]],
        ["Sin fecha entrega", stats.delivery_compliance["no_date"]],
    ])

    ws1 = wb.create_sheet("Carga maquinas")
    _write_sheet(ws1, ["Máquina", "Días hábiles", "Kg planificados"], [
        [r["machine"], r["working_days"], r.get("kg", 0)] for r in stats.machine_load
    ])

    ws2 = wb.create_sheet("Por material")
    _write_sheet(ws2, ["Material", "Órdenes", "Kg planificados"], [
        [r["material"], r["count"], r.get("kg", 0)] for r in stats.by_material
    ])

    ws3 = wb.create_sheet("Por cliente")
    _write_sheet(ws3, ["Cliente", "Órdenes", "Kg planificados"], [
        [r["customer"], r["count"], r.get("kg", 0)] for r in stats.by_customer
    ])

    return _wb_response(wb, f"dashboard_{date.today().isoformat()}.xlsx")
