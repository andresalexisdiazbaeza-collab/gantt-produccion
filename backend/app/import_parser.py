from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any, List, Optional

from openpyxl import load_workbook


def _to_float(val: Any) -> Optional[float]:
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _to_date(val: Any) -> Optional[date]:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def _is_valid_order_number(val: Any) -> bool:
    if val is None:
        return False
    s = str(val).strip()
    return s.isdigit() and len(s) >= 6


def parse_nuevo_formato(file_bytes: bytes) -> List[dict[str, Any]]:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows: List[dict[str, Any]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 4:
            continue
        order_number = row[3]
        if not _is_valid_order_number(order_number):
            continue

        data = {
            "raw_material": str(row[0]).strip() if row[0] is not None else None,
            "titulo": str(row[1]).strip() if row[1] is not None else None,
            "customer": str(row[2]).strip() if row[2] is not None else None,
            "order_number": str(int(float(order_number))) if order_number else None,
            "order_type": str(row[4]).strip() if len(row) > 4 and row[4] is not None else None,
            "braiding": str(row[5]).strip() if len(row) > 5 and row[5] is not None else None,
            "knot": _to_float(row[6]) if len(row) > 6 else None,
            "model": str(row[7]).strip() if len(row) > 7 and row[7] is not None else None,
            "matriz_mm": _to_float(row[8]) if len(row) > 8 else None,
            "measure": str(row[9]).strip() if len(row) > 9 and row[9] is not None else None,
            "meshes": _to_float(row[10]) if len(row) > 10 else None,
            "color": str(row[11]).strip() if len(row) > 11 and row[11] is not None else None,
            "treatment": str(row[12]).strip() if len(row) > 12 and row[12] is not None else None,
            "pieces": _to_float(row[13]) if len(row) > 13 else None,
            "piece_length": _to_float(row[14]) if len(row) > 14 else None,
            "kg_totales": _to_float(row[15]) if len(row) > 15 else None,
            "delivered": _to_float(row[16]) if len(row) > 16 else None,
            "delivery_date": _to_date(row[17]) if len(row) > 17 else None,
            "source_status": str(row[18]).strip() if len(row) > 18 and row[18] is not None else None,
        }
        rows.append(data)

    return rows


def _normalize_treatment(val: Any) -> Optional[str]:
    if val is None or val == "":
        return "-"
    s = str(val).strip()
    if not s or s.lower() in ("column5", "none"):
        return "-"
    return s


def parse_production_gantt(file_bytes: bytes) -> List[dict[str, Any]]:
    """Parse Production gantt2.xlsx — header row 6, data from row 7."""
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows: List[dict[str, Any]] = []

    for row_idx in range(7, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=c).value for c in range(1, 29)]
        if not row or len(row) < 4:
            continue
        order_number = row[3]
        if not _is_valid_order_number(order_number):
            continue

        start_date = _to_date(row[22]) if len(row) > 22 else None
        if start_date and start_date.year < 2000:
            start_date = None

        machine_raw = row[21] if len(row) > 21 else None
        machine_name = None
        if machine_raw is not None:
            try:
                machine_name = str(int(float(machine_raw)))
            except (TypeError, ValueError):
                machine_name = str(machine_raw).strip()

        data = {
            "raw_material": str(row[0]).strip() if row[0] is not None else None,
            "titulo": str(row[1]).strip() if row[1] is not None else None,
            "customer": str(row[2]).strip() if row[2] is not None else None,
            "order_number": str(int(float(order_number))),
            "order_type": str(row[4]).strip() if len(row) > 4 and row[4] is not None else None,
            "braiding": str(row[5]).strip() if len(row) > 5 and row[5] is not None else None,
            "knot": _to_float(row[6]) if len(row) > 6 else None,
            "model": str(row[7]).strip() if len(row) > 7 and row[7] is not None else None,
            "matriz_mm": _to_float(row[8]) if len(row) > 8 else None,
            "measure": str(row[9]).strip() if len(row) > 9 and row[9] is not None else None,
            "meshes": _to_float(row[10]) if len(row) > 10 else None,
            "color": str(row[11]).strip() if len(row) > 11 and row[11] is not None else None,
            "treatment": _normalize_treatment(row[12]) if len(row) > 12 else "-",
            "pieces": _to_float(row[13]) if len(row) > 13 else None,
            "piece_length": _to_float(row[14]) if len(row) > 14 else None,
            "kg_totales": _to_float(row[15]) if len(row) > 15 else None,
            "delivered": _to_float(row[16]) if len(row) > 16 else None,
            "delivery_date": _to_date(row[17]) if len(row) > 17 else None,
            "source_status": str(row[18]).strip() if len(row) > 18 and row[18] is not None else None,
            "machine_name": machine_name,
            "start_date": start_date,
            "comments": str(row[27]).strip() if len(row) > 27 and row[27] is not None else None,
        }
        rows.append(data)

    return rows
