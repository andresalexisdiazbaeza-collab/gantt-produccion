from __future__ import annotations

import os
import tempfile
import zipfile
from datetime import date, datetime
from typing import Any, Optional

from openpyxl import load_workbook

# Strict OOXML (Excel/Numbers export) → transitional namespaces openpyxl understands
_STRICT_TO_TRANS = (
    (
        b"http://purl.oclc.org/ooxml/spreadsheetml/main",
        b"http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    ),
    (
        b"http://purl.oclc.org/ooxml/officeDocument/relationships",
        b"http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    ),
    (
        b"http://purl.oclc.org/ooxml/drawingml/main",
        b"http://schemas.openxmlformats.org/drawingml/2006/main",
    ),
    (
        b"http://purl.oclc.org/ooxml/officeDocument/relationships/officeDocument",
        b"http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument",
    ),
    (
        b"http://purl.oclc.org/ooxml/officeDocument/relationships/worksheet",
        b"http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet",
    ),
    (
        b"http://purl.oclc.org/ooxml/officeDocument/relationships/sharedStrings",
        b"http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings",
    ),
    (
        b"http://purl.oclc.org/ooxml/officeDocument/relationships/styles",
        b"http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles",
    ),
    (
        b"http://purl.oclc.org/ooxml/officeDocument/relationships/theme",
        b"http://schemas.openxmlformats.org/officeDocument/2006/relationships/theme",
    ),
    (
        b"http://purl.oclc.org/ooxml/officeDocument/relationships/comments",
        b"http://schemas.openxmlformats.org/officeDocument/2006/relationships/comments",
    ),
    (
        b"http://purl.oclc.org/ooxml/officeDocument/relationships/drawing",
        b"http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing",
    ),
    (
        b"http://purl.oclc.org/ooxml/officeDocument/relationships/externalLink",
        b"http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLink",
    ),
)


def _needs_strict_ooxml_repair(file_path: str) -> bool:
    try:
        with zipfile.ZipFile(file_path) as zf:
            data = zf.read("xl/workbook.xml")
        return b"purl.oclc.org/ooxml" in data
    except Exception:
        return False


def _repair_strict_ooxml(file_path: str) -> str:
    """Rewrite strict OOXML namespaces so openpyxl can load the workbook."""
    out = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    out.close()
    with zipfile.ZipFile(file_path, "r") as zin, zipfile.ZipFile(out.name, "w", zipfile.ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            raw = zin.read(info.filename)
            name = info.filename
            if name.endswith((".xml", ".rels")) or name == "[Content_Types].xml":
                for old, new in _STRICT_TO_TRANS:
                    raw = raw.replace(old, new)
            zout.writestr(info, raw)
    return out.name


def _load_workbook(file_path: str):
    repaired: Optional[str] = None
    try:
        return load_workbook(file_path, data_only=True, read_only=True), None
    except TypeError:
        if not _needs_strict_ooxml_repair(file_path):
            raise
        repaired = _repair_strict_ooxml(file_path)
        return load_workbook(repaired, data_only=True, read_only=True), repaired


def _cell(row: tuple, idx: int) -> Any:
    if idx >= len(row):
        return None
    return row[idx]


def _as_str(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    text = str(value).strip()
    return text or None


def _as_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _as_date(value: Any) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _parse_quantity(pcs_label: Optional[str]) -> float:
    if not pcs_label:
        return 1.0
    text = pcs_label.lower().replace("pcs", "").strip()
    if " of " in text:
        return 1.0
    try:
        return float(text.replace(",", "."))
    except ValueError:
        return 1.0


def parse_confection_orders(file_path: str) -> list[dict[str, Any]]:
    wb, repaired = _load_workbook(file_path)
    try:
        if "ORDERS" not in wb.sheetnames:
            raise ValueError("Hoja ORDERS no encontrada")
        ws = wb["ORDERS"]
        rows = ws.iter_rows(min_row=4, values_only=True)
        items: list[dict[str, Any]] = []
        for row in rows:
            po = _as_str(_cell(row, 1))
            if not po:
                continue
            try:
                int(str(po).split()[0])
            except ValueError:
                continue
            pcs_label = _as_str(_cell(row, 0))
            netting = _cell(row, 17)
            netting_date = _as_date(netting)
            items.append(
                {
                    "pcs_label": pcs_label,
                    "quantity": _parse_quantity(pcs_label),
                    "po_number": po,
                    "purchase_order": _as_str(_cell(row, 2)),
                    "id_code": _as_str(_cell(row, 3)),
                    "customer": _as_str(_cell(row, 4)),
                    "tag_numbers": _as_str(_cell(row, 5)),
                    "circumference": _as_str(_cell(row, 6)),
                    "height": _as_str(_cell(row, 7)),
                    "cage_type": _as_str(_cell(row, 8)),
                    "mesh_mm": _as_str(_cell(row, 9)),
                    "twine_size": _as_str(_cell(row, 10)),
                    "color": _as_str(_cell(row, 11)),
                    "product_type": _as_str(_cell(row, 12)),
                    "received_date": _as_date(_cell(row, 13)),
                    "payment_terms": _as_str(_cell(row, 14)),
                    "requested_delivery_text": _as_str(_cell(row, 15)),
                    "delivery_offered": _as_date(_cell(row, 16)),
                    "netting_ready_date": netting_date,
                    "netting_status": None if netting_date else _as_str(netting),
                    "kg_cage": _as_float(_cell(row, 18)),
                    "netting_m2": _as_float(_cell(row, 19)),
                    "netting_kg": _as_float(_cell(row, 20)),
                    "total_hours": _as_float(_cell(row, 21)),
                    "coating_hours": _as_float(_cell(row, 22)),
                    "status": "activa",
                }
            )
        return items
    finally:
        wb.close()
        if repaired and os.path.exists(repaired):
            os.unlink(repaired)


def parse_confection_finished(file_path: str, sheet_name: str = "Finished 2024-2026") -> list[dict[str, Any]]:
    wb, repaired = _load_workbook(file_path)
    try:
        target = sheet_name if sheet_name in wb.sheetnames else next(
            (n for n in wb.sheetnames if n.lower().startswith("finished") and "2024" in n),
            None,
        )
        if not target:
            return []
        ws = wb[target]
        rows = ws.iter_rows(min_row=2, values_only=True)
        items: list[dict[str, Any]] = []
        for row in rows:
            po = _as_str(_cell(row, 1))
            if not po:
                continue
            pcs_label = _as_str(_cell(row, 0))
            finished_at = _as_date(_cell(row, 25))
            items.append(
                {
                    "pcs_label": pcs_label,
                    "quantity": _parse_quantity(pcs_label),
                    "po_number": po,
                    "purchase_order": _as_str(_cell(row, 2)),
                    "id_code": _as_str(_cell(row, 3)),
                    "customer": _as_str(_cell(row, 4)),
                    "tag_numbers": _as_str(_cell(row, 5)),
                    "circumference": _as_str(_cell(row, 6)),
                    "height": _as_str(_cell(row, 7)),
                    "cage_type": _as_str(_cell(row, 8)),
                    "mesh_mm": _as_str(_cell(row, 9)),
                    "twine_size": _as_str(_cell(row, 10)),
                    "color": _as_str(_cell(row, 11)),
                    "product_type": _as_str(_cell(row, 12)),
                    "received_date": _as_date(_cell(row, 13)),
                    "payment_terms": _as_str(_cell(row, 14)),
                    "requested_delivery_text": _as_str(_cell(row, 15)),
                    "delivery_offered": _as_date(_cell(row, 16)),
                    "kg_cage": _as_float(_cell(row, 17)),
                    "netting_m2": _as_float(_cell(row, 18)),
                    "netting_kg": _as_float(_cell(row, 19)),
                    "total_hours": _as_float(_cell(row, 20)),
                    "coating_hours": _as_float(_cell(row, 21)),
                    "real_hours": _as_float(_cell(row, 22)),
                    "pct_done": 100.0,
                    "status": "terminada",
                    "completed_at": finished_at,
                }
            )
        return items
    finally:
        wb.close()
        if repaired and os.path.exists(repaired):
            os.unlink(repaired)
