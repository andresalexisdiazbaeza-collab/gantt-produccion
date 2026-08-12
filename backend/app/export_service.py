from __future__ import annotations

import zipfile
from datetime import date, datetime
from io import BytesIO
from typing import Any, Iterable, List, Optional, Sequence

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.chart import BarChart, PieChart, Reference  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
import matplotlib.dates as mdates  # noqa: E402
from openpyxl.drawing.image import Image as XLImage  # noqa: E402
from reportlab.lib import colors  # noqa: E402
from reportlab.lib.pagesizes import A4, landscape  # noqa: E402
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet  # noqa: E402
from reportlab.lib.units import cm  # noqa: E402
from reportlab.platypus import Image, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle  # noqa: E402
from sqlalchemy.orm import Session  # noqa: E402

from .export_i18n import ExportTranslator, get_export_translator
from .models import ItemStatus, MachineConfig, MaterialConfig, ProductionItem
from .optimize_service import build_optimization_preview
from .routers.dashboard import get_dashboard_stats
from .services import item_to_dict

APP_TITLE = "Gantt Producción"
HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1E3A5F")
ALT_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
THIN = Side(style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ORDER_HEADERS = [
    "Orden", "Cliente", "Material", "Título", "Color", "Matriz mm", "Mallas", "Piezas",
    "Long. pieza", "Kg", "Entregado", "Fecha entrega", "Máquina", "Fecha inicio",
    "Fecha fin", "Días hábiles", "Total m", "M terminados", "M pendientes", "Shrinking", "Estado",
    "Cumplimiento", "Días atraso", "Días margen", "Comentarios", "Notas",
]


def _today_stamp() -> str:
    return date.today().isoformat()


def _order_row(item: dict) -> List[Any]:
    return [
        item.get("order_number"),
        item.get("customer"),
        item.get("raw_material"),
        item.get("titulo"),
        item.get("color"),
        item.get("matriz_mm"),
        item.get("meshes"),
        item.get("pieces"),
        item.get("piece_length"),
        item.get("kg_totales"),
        item.get("delivered"),
        item.get("delivery_date"),
        item.get("machine_name"),
        item.get("start_date"),
        item.get("finish_date"),
        item.get("working_days"),
        item.get("total_length"),
        item.get("meters_produced"),
        item.get("remaining_length"),
        item.get("shrinking"),
        item.get("status"),
        _compliance_label(item),
        item.get("days_late") or 0,
        item.get("days_margin") or 0,
        item.get("comments"),
        item.get("notes"),
    ]


def _compliance_label(item: dict, tr: Optional[ExportTranslator] = None) -> str:
    if tr:
        return tr.compliance_label(item)
    status = item.get("delivery_status")
    if status == "late" or item.get("is_late"):
        return f"Atrasada +{item.get('days_late', 0)} d"
    if status == "on_time":
        margin = item.get("days_margin") or 0
        return f"A tiempo (−{margin} d)" if margin else "A tiempo"
    if status == "no_date":
        return "Sin fecha entrega"
    if status == "pending":
        return "Sin fecha fin"
    return ""


def _machine_sort_key(name: str) -> tuple:
    try:
        return (0, int(str(name).strip()))
    except ValueError:
        return (1, str(name).lower())


def _group_by_machine(items: List[dict], tr: ExportTranslator) -> List[tuple[str, List[dict]]]:
    buckets: dict[str, List[dict]] = {}
    for item in items:
        machine = item.get("machine_name") or tr.t("unassigned")
        buckets.setdefault(machine, []).append(item)
    groups: List[tuple[str, List[dict]]] = []
    for machine in sorted(buckets.keys(), key=_machine_sort_key):
        group = sorted(buckets[machine], key=lambda x: str(x.get("start_date") or ""))
        groups.append((machine, group))
    return groups


def _safe_sheet_name(name: str, prefix: str = "") -> str:
    cleaned = "".join(c for c in str(name) if c.isalnum() or c in " -_").strip() or "M"
    return (f"{prefix}{cleaned}"[:31]).strip()


def _write_table(ws, headers: Sequence[str], rows: Sequence[Sequence[Any]], start_row: int = 1) -> int:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r_idx, row in enumerate(rows, start_row + 1):
        fill = ALT_FILL if (r_idx - start_row) % 2 == 0 else None
        for c_idx, val in enumerate(row, 1):
            if isinstance(val, (date, datetime)):
                val = val.isoformat() if val else ""
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = BORDER
            if fill:
                cell.fill = fill
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        max_len = len(str(headers[col - 1]))
        for row in ws.iter_rows(min_row=start_row + 1, max_row=start_row + len(rows), min_col=col, max_col=col):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, min(len(str(cell.value)), 40))
        ws.column_dimensions[letter].width = max(10, min(max_len + 2, 36))
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    return start_row + len(rows)


def _add_bar_chart(ws, title: str, data_end_row: int, categories_col: int, values_col: int, anchor: str, height: int = 12):
    if data_end_row < 2:
        return
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.style = 10
    chart.y_axis.title = "Valor"
    chart.height = height
    chart.width = 22
    cats = Reference(ws, min_col=categories_col, min_row=2, max_row=data_end_row)
    vals = Reference(ws, min_col=values_col, min_row=1, max_row=data_end_row)
    chart.add_data(vals, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor)


def _add_pie_chart(ws, title: str, data_end_row: int, labels_col: int, values_col: int, anchor: str):
    if data_end_row < 2:
        return
    chart = PieChart()
    chart.title = title
    chart.height = 12
    chart.width = 16
    labels = Reference(ws, min_col=labels_col, min_row=2, max_row=data_end_row)
    data = Reference(ws, min_col=values_col, min_row=2, max_row=data_end_row)
    chart.add_data(data)
    chart.set_categories(labels)
    ws.add_chart(chart, anchor)


def _cover_sheet(wb: Workbook, modules: Iterable[str]) -> None:
    ws = wb.active
    ws.title = "Portada"
    ws["A1"] = APP_TITLE
    ws["A1"].font = Font(bold=True, size=18, color="1E3A5F")
    ws["A2"] = f"Informe generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A3"] = f"Módulos incluidos: {', '.join(modules)}"
    ws.column_dimensions["A"].width = 80


def _fetch_orders(db: Session, status: Optional[str] = None) -> List[dict]:
    q = db.query(ProductionItem)
    if status:
        q = q.filter(ProductionItem.status == status)
    items = q.order_by(ProductionItem.order_number).all()
    return [item_to_dict(i) for i in items]


def build_orders_workbook(db: Session, status: Optional[str] = None, title: str = "Ordenes", lang: str = "es") -> Workbook:
    rows = [_order_row(i) for i in _fetch_orders(db, status)]
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    end = _write_table(ws, ORDER_HEADERS, rows)
    by_machine: dict[str, float] = {}
    for item in _fetch_orders(db, status):
        name = item.get("machine_name") or "Sin asignar"
        by_machine[name] = by_machine.get(name, 0) + float(item.get("working_days") or 0)
    if by_machine:
        start = end + 3
        ws.cell(row=start, column=1, value="Resumen por máquina").font = TITLE_FONT
        summary_headers = ["Máquina", "Días hábiles"]
        summary_rows = [[k, round(v, 2)] for k, v in sorted(by_machine.items())]
        summary_end = _write_table(ws, summary_headers, summary_rows, start_row=start + 1)
        _add_bar_chart(ws, "Carga por máquina (días)", summary_end, 1, 2, f"E{summary_end + 3}")
    return wb


def _parse_date(val: Any) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, date):
        return val
    if isinstance(val, datetime):
        return val.date()
    try:
        return date.fromisoformat(str(val)[:10])
    except ValueError:
        return None


def _scheduled_active_items(db: Session) -> List[dict]:
    items = _fetch_orders(db, ItemStatus.ACTIVA.value)
    return [i for i in items if i.get("start_date") and i.get("finish_date")]


def _build_gantt_timeline_figure(items: List[dict], tr: ExportTranslator, machine_filter: Optional[str] = None):
    if machine_filter:
        items = [
            i for i in items
            if (i.get("machine_name") or tr.t("unassigned")) == machine_filter
        ]
    groups = _group_by_machine(items, tr)
    if not groups:
        return None

    row_count = sum(1 + len(group) + 1 for _, group in groups)
    fig_height = max(4.5, min(30, row_count * 0.55))
    fig, ax = plt.subplots(figsize=(14, fig_height))

    y_pos = 0
    y_ticks: List[float] = []
    y_labels: List[str] = []
    min_date: Optional[date] = None
    max_date: Optional[date] = None

    for machine, group_items in groups:
        ax.axhspan(y_pos - 0.42, y_pos + 0.42, color="#1E3A5F", alpha=0.08, zorder=0)
        y_ticks.append(y_pos)
        y_labels.append(f"▌ {tr.t('machine_section', name=machine)} ({len(group_items)})")
        y_pos += 1

        for item in group_items:
            start = _parse_date(item["start_date"])
            finish = _parse_date(item["finish_date"])
            if not start or not finish:
                continue
            min_date = start if min_date is None else min(min_date, start)
            max_date = finish if max_date is None else max(max_date, finish)
            duration = max(1, (finish - start).days + 1)
            late = bool(item.get("is_late"))
            color = "#ef4444" if late else "#3b82f6"
            start_num = mdates.date2num(start)
            ax.barh(y_pos, duration, left=start_num, height=0.68, color=color, edgecolor="white", linewidth=0.5, zorder=2)

            order = str(item.get("order_number") or "")
            customer = (item.get("customer") or "")[:16]
            bar_label = f"{order} · {customer}" if customer else order
            ax.text(
                start_num + 0.35,
                y_pos,
                bar_label[:32],
                va="center",
                fontsize=7,
                color="white",
                fontweight="bold",
                zorder=3,
            )
            y_ticks.append(y_pos)
            y_labels.append(f"   {order}")
            y_pos += 1

        y_pos += 0.55

    ax.set_yticks(y_ticks)
    ax.set_yticklabels(y_labels, fontsize=7)
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%d/%m"))
    ax.xaxis.set_major_locator(mdates.WeekdayLocator(interval=1))
    title_suffix = f" — {tr.t('machine_section', name=machine_filter)}" if machine_filter else ""
    ax.set_title(
        f"{tr.t('app_title')} — {tr.t('gantt_chart_title')}{title_suffix}",
        fontsize=13,
        color="#1E3A5F",
        fontweight="bold",
        pad=12,
    )
    ax.set_xlabel(tr.t("gantt_calendar"), fontsize=10, color="#475569")
    if min_date and max_date:
        pad = mdates.date2num(min_date) - 1
        ax.set_xlim(pad, mdates.date2num(max_date) + 2)
    ax.invert_yaxis()
    ax.grid(axis="x", alpha=0.25, linestyle="--")
    fig.autofmt_xdate(rotation=30)
    fig.tight_layout()
    return fig


def _fig_to_xlsx_image(fig, width: int = 960, height: int = 540) -> XLImage:
    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    buf.seek(0)
    img = XLImage(buf)
    img.width = width
    img.height = height
    return img


def _gantt_table_headers(tr: ExportTranslator) -> List[str]:
    return [
        tr.t("machine"), tr.t("order"), tr.t("customer"), tr.t("start"),
        tr.t("finish"), tr.t("days"), tr.t("delivery"), tr.t("compliance"),
    ]


def _gantt_table_row(item: dict, machine: str, tr: ExportTranslator) -> List[Any]:
    return [
        machine,
        item.get("order_number"),
        item.get("customer"),
        item.get("start_date"),
        item.get("finish_date"),
        item.get("working_days"),
        item.get("delivery_date"),
        _compliance_label(item, tr),
    ]


def _write_machine_summary_table(ws, scheduled: List[dict], tr: ExportTranslator, start_row: int = 1) -> int:
    groups = _group_by_machine(scheduled, tr)
    headers = [
        tr.t("machine"), tr.t("total_orders"), tr.t("working_days"),
        tr.t("late_orders"), tr.t("on_time"),
    ]
    rows: List[List[Any]] = []
    for machine, group in groups:
        late = sum(1 for i in group if i.get("is_late"))
        on_time = sum(1 for i in group if i.get("delivery_status") == "on_time")
        days = round(sum(float(i.get("working_days") or 0) for i in group), 2)
        rows.append([machine, len(group), days, late, on_time])
    return _write_table(ws, headers, rows, start_row=start_row)


def _add_machine_gantt_sheets(wb: Workbook, scheduled: List[dict], tr: ExportTranslator) -> None:
    groups = _group_by_machine(scheduled, tr)
    for machine, group in groups:
        sheet_name = _safe_sheet_name(machine, "M ")
        ws = wb.create_sheet(sheet_name)
        ws["A1"] = tr.t("machine_section", name=machine)
        ws["A1"].font = TITLE_FONT
        ws["A2"] = (
            f"{tr.t('gantt_generated')}: {datetime.now().strftime('%d/%m/%Y %H:%M')} · "
            f"{tr.t('machine_orders', count=len(group))}"
        )
        ws.merge_cells("A2:H2")

        rows = [_gantt_table_row(item, machine, tr) for item in group]
        table_end = _write_table(ws, _gantt_table_headers(tr), rows, start_row=4)

        fig = _build_gantt_timeline_figure(group, tr, machine_filter=machine)
        if fig:
            img_height = min(620, max(240, len(group) * 34 + 80))
            ws.add_image(_fig_to_xlsx_image(fig, width=980, height=img_height), f"A{table_end + 3}")

        legend_row = table_end + 3 + max(12, len(group) // 2 + 8)
        ws.cell(row=legend_row, column=1, value=f"{tr.t('legend_title')}:").font = Font(bold=True)
        ws.cell(row=legend_row + 1, column=1, value=tr.t("legend_on_time"))
        ws.cell(row=legend_row + 2, column=1, value=tr.t("legend_late"))


def _add_gantt_sheet(wb: Workbook, db: Session, lang: str = "es") -> None:
    tr = get_export_translator(lang)
    scheduled = _scheduled_active_items(db)

    ws = wb.create_sheet(tr.t("sheet_summary")[:31])
    ws["A1"] = f"{tr.t('app_title')} — {tr.t('gantt_plan_title')}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        f"{tr.t('gantt_generated')}: {datetime.now().strftime('%d/%m/%Y %H:%M')} · "
        f"{len(scheduled)} {tr.t('gantt_scheduled_orders')}"
    )
    ws.merge_cells("A2:H2")

    if not scheduled:
        ws["A4"] = tr.t("gantt_no_schedule")
        return

    ws["A4"] = tr.t("summary_by_machine")
    ws["A4"].font = Font(bold=True, size=11, color="1E3A5F")
    summary_end = _write_machine_summary_table(ws, scheduled, tr, start_row=5)

    fig = _build_gantt_timeline_figure(scheduled, tr)
    if fig:
        img_height = min(780, max(320, len(scheduled) * 22 + 120))
        ws.add_image(_fig_to_xlsx_image(fig, width=980, height=img_height), f"A{summary_end + 3}")

    legend_row = summary_end + 3 + max(16, len(scheduled) // 2 + 10)
    ws.cell(row=legend_row, column=1, value=f"{tr.t('legend_title')}:").font = Font(bold=True)
    ws.cell(row=legend_row + 1, column=1, value=tr.t("legend_on_time"))
    ws.cell(row=legend_row + 2, column=1, value=tr.t("legend_late"))
    ws.cell(row=legend_row + 3, column=1, value=tr.t("gantt_legend"))

    _add_machine_gantt_sheets(wb, scheduled, tr)


def build_active_orders_gantt_workbook(db: Session, lang: str = "es") -> Workbook:
    tr = get_export_translator(lang)
    wb = build_orders_workbook(db, ItemStatus.ACTIVA.value, tr.t("sheet_orders"))
    _add_gantt_sheet(wb, db, lang)
    return wb


def build_active_orders_gantt_pdf(db: Session, lang: str = "es") -> bytes:
    tr = get_export_translator(lang)
    items = _fetch_orders(db, ItemStatus.ACTIVA.value)
    scheduled = _scheduled_active_items(db)
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1 * cm, rightMargin=1 * cm)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="Subtitle", fontSize=10, textColor=colors.grey))

    story: list[Any] = [
        _pdf_section_title(f"{tr.t('app_title')} — {tr.t('active_orders_gantt')}", styles),
        Paragraph(
            f"{tr.t('records')}: {len(items)} · {tr.t('scheduled')}: {len(scheduled)} · "
            f"{tr.t('gantt_generated')}: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            styles["Subtitle"],
        ),
        Spacer(1, 0.3 * cm),
    ]

    headers = [
        tr.t("order"), tr.t("customer"), tr.t("material"), tr.t("machine"),
        tr.t("start"), tr.t("finish"), tr.t("delivery"), tr.t("compliance"), tr.t("kg"),
    ]
    rows = [[
        i.get("order_number"), i.get("customer"), i.get("raw_material"),
        i.get("machine_name"), i.get("start_date"), i.get("finish_date"),
        i.get("delivery_date"), _compliance_label(i, tr), i.get("kg_totales"),
    ] for i in items[:150]]
    story.append(_pdf_table([headers] + rows))
    if len(items) > 150:
        story.append(Paragraph(tr.t("more_in_excel", count=len(items) - 150), styles["Italic"]))

    if scheduled:
        groups = _group_by_machine(scheduled, tr)
        story += [
            PageBreak(),
            _pdf_section_title(tr.t("summary_by_machine"), styles),
            Spacer(1, 0.2 * cm),
        ]
        summary_headers = [
            tr.t("machine"), tr.t("total_orders"), tr.t("working_days"),
            tr.t("late_orders"), tr.t("on_time"),
        ]
        summary_rows = []
        for machine, group in groups:
            late = sum(1 for i in group if i.get("is_late"))
            on_time = sum(1 for i in group if i.get("delivery_status") == "on_time")
            days = round(sum(float(i.get("working_days") or 0) for i in group), 2)
            summary_rows.append([machine, len(group), days, late, on_time])
        story.append(_pdf_table([summary_headers] + summary_rows))

        story += [
            PageBreak(),
            _pdf_section_title(tr.t("gantt_chart_title"), styles),
            Paragraph(tr.t("gantt_legend"), styles["Subtitle"]),
            Spacer(1, 0.3 * cm),
        ]
        overview_fig = _build_gantt_timeline_figure(scheduled, tr)
        if overview_fig:
            story.append(_fig_to_image(overview_fig, width=24 * cm))

        for machine, group in groups:
            fig = _build_gantt_timeline_figure(group, tr, machine_filter=machine)
            if not fig:
                continue
            story += [
                PageBreak(),
                _pdf_section_title(tr.t("machine_section", name=machine), styles),
                Paragraph(tr.t("machine_orders", count=len(group)), styles["Subtitle"]),
                Spacer(1, 0.25 * cm),
                _pdf_table([_gantt_table_headers(tr)] + [_gantt_table_row(i, machine, tr) for i in group]),
                Spacer(1, 0.3 * cm),
                _fig_to_image(fig, width=24 * cm),
            ]

        on_time = sum(1 for i in items if i.get("delivery_status") == "on_time")
        late = sum(1 for i in items if i.get("is_late"))
        no_date = sum(1 for i in items if i.get("delivery_status") == "no_date")
        pending = sum(1 for i in items if i.get("delivery_status") == "pending")
        labels_vals = [
            (tr.t("on_time"), on_time),
            (tr.t("late"), late),
            (tr.t("no_date"), no_date),
            (tr.t("pending_finish"), pending),
        ]
        filtered = [(label, val) for label, val in labels_vals if val > 0]
        if filtered:
            pie_labels, pie_values = zip(*filtered)
            story += [
                Spacer(1, 0.4 * cm),
                _fig_to_image(_build_chart_figure(
                    list(pie_labels), list(pie_values),
                    tr.t("delivery_compliance"),
                    "pie",
                ), width=14 * cm),
            ]

    doc.build(story)
    return buf.getvalue()


def build_gantt_workbook(db: Session, lang: str = "es") -> Workbook:
    tr = get_export_translator(lang)
    scheduled = _scheduled_active_items(db)
    wb = Workbook()
    ws = wb.active
    ws.title = tr.t("sheet_summary")[:31]
    ws["A1"] = f"{tr.t('app_title')} — {tr.t('gantt_plan_title')}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        f"{tr.t('gantt_generated')}: {datetime.now().strftime('%d/%m/%Y %H:%M')} · "
        f"{len(scheduled)} {tr.t('gantt_scheduled_orders')}"
    )
    ws.merge_cells("A2:H2")
    if not scheduled:
        ws["A4"] = tr.t("gantt_no_schedule")
        return wb
    ws["A4"] = tr.t("summary_by_machine")
    ws["A4"].font = Font(bold=True, size=11, color="1E3A5F")
    summary_end = _write_machine_summary_table(ws, scheduled, tr, start_row=5)
    fig = _build_gantt_timeline_figure(scheduled, tr)
    if fig:
        img_height = min(780, max(320, len(scheduled) * 22 + 120))
        ws.add_image(_fig_to_xlsx_image(fig, width=980, height=img_height), f"A{summary_end + 3}")
    _add_machine_gantt_sheets(wb, scheduled, tr)
    return wb


def build_gantt_pdf(db: Session, lang: str = "es") -> bytes:
    tr = get_export_translator(lang)
    scheduled = _scheduled_active_items(db)
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1 * cm, rightMargin=1 * cm)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="Subtitle", fontSize=10, textColor=colors.grey))

    story: list[Any] = [
        _pdf_section_title(f"{tr.t('app_title')} — {tr.t('gantt_chart_title')}", styles),
        Paragraph(
            f"{tr.t('scheduled')}: {len(scheduled)} · "
            f"{tr.t('gantt_generated')}: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            styles["Subtitle"],
        ),
        Spacer(1, 0.3 * cm),
    ]

    if not scheduled:
        story.append(Paragraph(tr.t("gantt_no_schedule"), styles["Normal"]))
        doc.build(story)
        return buf.getvalue()

    groups = _group_by_machine(scheduled, tr)
    summary_headers = [
        tr.t("machine"), tr.t("total_orders"), tr.t("working_days"),
        tr.t("late_orders"), tr.t("on_time"),
    ]
    summary_rows = []
    for machine, group in groups:
        late = sum(1 for i in group if i.get("is_late"))
        on_time = sum(1 for i in group if i.get("delivery_status") == "on_time")
        days = round(sum(float(i.get("working_days") or 0) for i in group), 2)
        summary_rows.append([machine, len(group), days, late, on_time])
    story.append(_pdf_table([summary_headers] + summary_rows))

    story += [
        PageBreak(),
        Paragraph(tr.t("gantt_legend"), styles["Subtitle"]),
        Spacer(1, 0.2 * cm),
    ]
    overview_fig = _build_gantt_timeline_figure(scheduled, tr)
    if overview_fig:
        story.append(_fig_to_image(overview_fig, width=24 * cm))

    for machine, group in groups:
        fig = _build_gantt_timeline_figure(group, tr, machine_filter=machine)
        if not fig:
            continue
        story += [
            PageBreak(),
            _pdf_section_title(tr.t("machine_section", name=machine), styles),
            Paragraph(tr.t("machine_orders", count=len(group)), styles["Subtitle"]),
            Spacer(1, 0.25 * cm),
            _pdf_table([_gantt_table_headers(tr)] + [_gantt_table_row(i, machine, tr) for i in group]),
            Spacer(1, 0.3 * cm),
            _fig_to_image(fig, width=24 * cm),
        ]

    doc.build(story)
    return buf.getvalue()


def build_dashboard_workbook(db: Session, lang: str = "es") -> Workbook:
    stats = get_dashboard_stats(db)
    wb = Workbook()

    ws0 = wb.active
    ws0.title = "KPIs"
    kpi_rows = [
        ["Órdenes activas", stats.active_count],
        ["Terminadas", stats.completed_count],
        ["Máquinas activas", stats.machines_active],
        ["Kg planificados", stats.total_planned_kg],
        ["Kg producidos", stats.total_produced_kg],
        ["Kg pendientes", stats.total_remaining_kg],
        ["Metros planificados", stats.total_planned_meters],
        ["A tiempo", stats.delivery_compliance["on_time"]],
        ["Tardíos", stats.delivery_compliance["late"]],
        ["Sin fecha entrega", stats.delivery_compliance["no_date"]],
    ]
    kpi_end = _write_table(ws0, ["Indicador", "Valor"], kpi_rows)

    ws1 = wb.create_sheet("Carga maquinas")
    m_rows = [[r["machine"], r["working_days"], r.get("kg", 0)] for r in stats.machine_load]
    m_end = _write_table(ws1, ["Máquina", "Días hábiles", "Kg planificados"], m_rows)
    _add_bar_chart(ws1, "Días hábiles por máquina", m_end, 1, 2, f"E{m_end + 2}")
    _add_bar_chart(ws1, "Kg planificados por máquina", m_end, 1, 3, f"E{m_end + 20}")

    ws2 = wb.create_sheet("Por material")
    mat_rows = [[r["material"], r["count"], r.get("kg", 0)] for r in stats.by_material]
    mat_end = _write_table(ws2, ["Material", "Órdenes", "Kg planificados"], mat_rows)
    _add_pie_chart(ws2, "Kg por material", mat_end, 1, 3, f"E{mat_end + 2}")

    ws3 = wb.create_sheet("Por cliente")
    cust_rows = [[r["customer"], r["count"], r.get("kg", 0)] for r in stats.by_customer]
    cust_end = _write_table(ws3, ["Cliente", "Órdenes", "Kg planificados"], cust_rows)
    _add_bar_chart(ws3, "Top clientes (kg)", cust_end, 1, 3, f"E{cust_end + 2}")

    ws4 = wb.create_sheet("Cumplimiento")
    comp_rows = [
        ["A tiempo", stats.delivery_compliance["on_time"]],
        ["Tardíos", stats.delivery_compliance["late"]],
        ["Sin fecha", stats.delivery_compliance["no_date"]],
    ]
    comp_end = _write_table(ws4, ["Estado", "Órdenes"], comp_rows)
    _add_pie_chart(ws4, "Cumplimiento entregas", comp_end, 1, 2, f"D{comp_end + 2}")

    ws0.cell(row=kpi_end + 3, column=1, value="Dashboard KPIs").font = TITLE_FONT
    return wb


def build_materials_workbook(db: Session, lang: str = "es") -> Workbook:
    materials = db.query(MaterialConfig).order_by(MaterialConfig.material).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Materiales"
    _write_table(ws, ["Material", "Shrinking", "Actualizado"], [
        [m.material, m.shrinking, m.updated_at.isoformat() if m.updated_at else ""]
        for m in materials
    ])
    return wb


def build_machines_workbook(db: Session, lang: str = "es") -> Workbook:
    machines = db.query(MachineConfig).order_by(MachineConfig.name).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Maquinas"
    _write_table(ws, ["Máquina", "m/turno", "Turnos/día", "Cambio (turnos)", "Activa"], [
        [m.name, m.mts_per_shift, m.shifts_per_day, m.changeover_shifts, "Sí" if m.active else "No"]
        for m in machines
    ])
    return wb


def build_optimize_workbook(db: Session, lang: str = "es") -> Workbook:
    preview = build_optimization_preview(db)
    wb = Workbook()
    ws = wb.active
    ws.title = "Metricas"
    cur = preview["current"]
    opt = preview["optimized"]
    imp = preview["improvement"]
    metric_rows = [
        ["Programadas", cur["scheduled_count"], opt["scheduled_count"]],
        ["A tiempo", cur["on_time"], opt["on_time"]],
        ["Tardías", cur["late"], opt["late"]],
        ["Sin fecha entrega", cur["no_delivery_date"], opt["no_delivery_date"]],
        ["Cambios setup", cur["total_changeovers"], opt["total_changeovers"]],
        ["Turnos setup", cur["total_setup_shifts"], opt["total_setup_shifts"]],
        ["Cambios ahorrados", "", imp["changeovers_saved"]],
        ["Turnos ahorrados", "", imp["setup_shifts_saved"]],
        ["Tardías reducidas", "", imp["late_reduced"]],
    ]
    _write_table(ws, ["Métrica", "Actual", "Optimizado"], metric_rows)

    for label, plans in [("Plan actual", preview["current_machines"]), ("Plan optimizado", preview["optimized_machines"])]:
        sheet = wb.create_sheet(label[:31])
        headers = [
            "Máquina", "Orden", "Cliente", "Título", "Color", "Inicio", "Fin",
            "Entrega", "Días", "Setup turnos", "Tardía", "Días tarde",
        ]
        rows: List[List[Any]] = []
        for plan in plans:
            for slot in plan["items"]:
                rows.append([
                    plan["machine_name"],
                    slot["order_number"],
                    slot.get("customer"),
                    slot.get("titulo"),
                    slot.get("color"),
                    slot.get("start_date"),
                    slot.get("finish_date"),
                    slot.get("delivery_date"),
                    slot.get("working_days"),
                    slot.get("setup_shifts"),
                    "Sí" if slot.get("is_late") else "No",
                    slot.get("days_late"),
                ])
        _write_table(sheet, headers, rows)
    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _pdf_table(data: List[List[Any]], col_widths: Optional[List[float]] = None) -> Table:
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    return table


def _fig_to_image(fig, width: float = 16 * cm) -> Image:
    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=140, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    buf.seek(0)
    return Image(buf, width=width, height=width * 0.55)


def _pdf_section_title(text: str, styles) -> Paragraph:
    return Paragraph(text, styles["Heading1"])


def _build_chart_figure(labels: List[str], values: List[float], title: str, kind: str = "bar"):
    fig, ax = plt.subplots(figsize=(8, 4))
    if kind == "pie":
        ax.pie(values, labels=labels, autopct="%1.0f%%", startangle=90)
    else:
        ax.bar(labels, values, color="#3b82f6")
        ax.tick_params(axis="x", rotation=30, labelsize=8)
    ax.set_title(title, fontsize=11, color="#1E3A5F")
    fig.tight_layout()
    return fig


def build_dashboard_pdf(db: Session, lang: str = "es") -> bytes:
    stats = get_dashboard_stats(db)
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1.5 * cm, rightMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="Subtitle", fontSize=10, textColor=colors.grey))
    story: list[Any] = [
        _pdf_section_title(f"{APP_TITLE} — Dashboard", styles),
        Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["Subtitle"]),
        Spacer(1, 0.4 * cm),
        _pdf_table([
            ["Indicador", "Valor"],
            ["Órdenes activas", stats.active_count],
            ["Terminadas", stats.completed_count],
            ["Kg planificados", stats.total_planned_kg],
            ["Kg producidos", stats.total_produced_kg],
            ["Kg pendientes", stats.total_remaining_kg],
        ], [8 * cm, 6 * cm]),
    ]
    if stats.machine_load:
        labels = [r["machine"] for r in stats.machine_load]
        values = [r["working_days"] for r in stats.machine_load]
        story += [Spacer(1, 0.5 * cm), _fig_to_image(_build_chart_figure(labels, values, "Carga por máquina (días hábiles)"))]
    if stats.by_material:
        labels = [r["material"] for r in stats.by_material[:8]]
        values = [r.get("kg", 0) for r in stats.by_material[:8]]
        story += [PageBreak(), _fig_to_image(_build_chart_figure(labels, values, "Kg por material", "pie"))]
    comp_labels = ["A tiempo", "Tardíos", "Sin fecha"]
    comp_values = [
        stats.delivery_compliance["on_time"],
        stats.delivery_compliance["late"],
        stats.delivery_compliance["no_date"],
    ]
    if any(comp_values):
        story += [Spacer(1, 0.5 * cm), _fig_to_image(_build_chart_figure(comp_labels, comp_values, "Cumplimiento entregas", "pie"))]
    doc.build(story)
    return buf.getvalue()


def build_orders_pdf(db: Session, status: Optional[str] = None, title: str = "Órdenes", lang: str = "es") -> bytes:
    items = _fetch_orders(db, status)
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1 * cm, rightMargin=1 * cm)
    styles = getSampleStyleSheet()
    story: list[Any] = [
        _pdf_section_title(f"{APP_TITLE} — {title}", styles),
        Paragraph(f"Registros: {len(items)} · {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]),
        Spacer(1, 0.3 * cm),
    ]
    headers = ["Orden", "Cliente", "Material", "Máquina", "Inicio", "Fin", "Kg", "Estado"]
    rows = [[
        i.get("order_number"), i.get("customer"), i.get("raw_material"), i.get("machine_name"),
        i.get("start_date"), i.get("finish_date"), i.get("kg_totales"), i.get("status"),
    ] for i in items[:200]]
    story.append(_pdf_table([headers] + rows))
    if len(items) > 200:
        story.append(Paragraph(f"... y {len(items) - 200} registros más (ver Excel completo)", styles["Italic"]))
    doc.build(story)
    return buf.getvalue()


def build_materials_pdf(db: Session, lang: str = "es") -> bytes:
    materials = db.query(MaterialConfig).order_by(MaterialConfig.material).all()
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    rows = [["Material", "Shrinking"]] + [[m.material, m.shrinking] for m in materials]
    doc.build([
        _pdf_section_title(f"{APP_TITLE} — Materiales", styles),
        Spacer(1, 0.3 * cm),
        _pdf_table(rows, [8 * cm, 4 * cm]),
    ])
    return buf.getvalue()


def build_machines_pdf(db: Session, lang: str = "es") -> bytes:
    machines = db.query(MachineConfig).order_by(MachineConfig.name).all()
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    rows = [["Máquina", "m/turno", "Turnos/día", "Cambio", "Activa"]] + [
        [m.name, m.mts_per_shift, m.shifts_per_day, m.changeover_shifts, "Sí" if m.active else "No"]
        for m in machines
    ]
    doc.build([
        _pdf_section_title(f"{APP_TITLE} — Máquinas", styles),
        Spacer(1, 0.3 * cm),
        _pdf_table(rows),
    ])
    return buf.getvalue()


def build_optimize_pdf(db: Session, lang: str = "es") -> bytes:
    preview = build_optimization_preview(db)
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    cur, opt = preview["current"], preview["optimized"]
    story: list[Any] = [
        _pdf_section_title(f"{APP_TITLE} — Optimización", styles),
        _pdf_table([
            ["Métrica", "Actual", "Optimizado"],
            ["Programadas", cur["scheduled_count"], opt["scheduled_count"]],
            ["A tiempo", cur["on_time"], opt["on_time"]],
            ["Tardías", cur["late"], opt["late"]],
            ["Cambios setup", cur["total_changeovers"], opt["total_changeovers"]],
        ]),
        Spacer(1, 0.4 * cm),
        _fig_to_image(_build_chart_figure(
            ["A tiempo", "Tardías", "Cambios"],
            [cur["on_time"], cur["late"], cur["total_changeovers"]],
            "Plan actual",
        )),
        Spacer(1, 0.3 * cm),
        _fig_to_image(_build_chart_figure(
            ["A tiempo", "Tardías", "Cambios"],
            [opt["on_time"], opt["late"], opt["total_changeovers"]],
            "Plan optimizado",
        )),
    ]
    doc.build(story)
    return buf.getvalue()


def _copy_sheet(source_wb: Workbook, target_wb: Workbook, new_title: str) -> None:
    src = source_wb.active if len(source_wb.sheetnames) == 1 else source_wb.worksheets[0]
    ws = target_wb.create_sheet(new_title[:31])
    for row in src.iter_rows():
        for cell in row:
            new = ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new.font = cell.font.copy()
                new.fill = cell.fill.copy()
                new.border = cell.border.copy()
                new.alignment = cell.alignment.copy()
    for col, dim in src.column_dimensions.items():
        ws.column_dimensions[col].width = dim.width
    if src.freeze_panes:
        ws.freeze_panes = src.freeze_panes


def _merge_workbooks(target: Workbook, source: Workbook, prefix: str = "") -> None:
    for sheet in source.worksheets:
        title = f"{prefix}{sheet.title}"[:31]
        if title in target.sheetnames:
            n = 2
            while f"{title[:28]}_{n}" in target.sheetnames:
                n += 1
            title = f"{title[:28]}_{n}"
        new_ws = target.create_sheet(title)
        for row in sheet.iter_rows():
            for cell in row:
                new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
        for col, dim in sheet.column_dimensions.items():
            new_ws.column_dimensions[col].width = dim.width


MODULE_BUILDERS = {
    "dashboard": ("Dashboard", build_dashboard_workbook, build_dashboard_pdf),
    "gantt": ("Gantt", build_gantt_workbook, build_gantt_pdf),
    "active_orders": ("Ordenes activas", build_active_orders_gantt_workbook, build_active_orders_gantt_pdf),
    "completed": (
        "Ordenes terminadas",
        lambda db, lang="es": build_orders_workbook(db, ItemStatus.TERMINADA.value, "Terminadas", lang),
        lambda db, lang="es": build_orders_pdf(db, ItemStatus.TERMINADA.value, "Órdenes terminadas", lang),
    ),
    "optimize": ("Optimizacion", build_optimize_workbook, build_optimize_pdf),
    "materials": ("Materiales", build_materials_workbook, build_materials_pdf),
    "machines": ("Maquinas", build_machines_workbook, build_machines_pdf),
}


def build_complete_workbook(db: Session, modules: Sequence[str], lang: str = "es") -> Workbook:
    wb = Workbook()
    _cover_sheet(wb, [MODULE_BUILDERS[m][0] for m in modules if m in MODULE_BUILDERS])
    for key in modules:
        if key not in MODULE_BUILDERS:
            continue
        label, xlsx_fn, _ = MODULE_BUILDERS[key]
        _merge_workbooks(wb, xlsx_fn(db, lang=lang), prefix=f"{label} - ")
    if "Portada" in wb.sheetnames and len(wb.sheetnames) > 1:
        portada = wb["Portada"]
        wb._sheets = [portada] + [s for s in wb.worksheets if s != portada]
    return wb


def build_complete_pdf(db: Session, modules: Sequence[str], lang: str = "es") -> bytes:
    parts: list[bytes] = []
    for key in modules:
        if key not in MODULE_BUILDERS:
            continue
        _, _, pdf_fn = MODULE_BUILDERS[key]
        parts.append(pdf_fn(db, lang=lang))
    if not parts:
        return b""
    if len(parts) == 1:
        return parts[0]
  # Concatenate PDFs by rebuilding one doc with all sections
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    story: list[Any] = [
        _pdf_section_title(f"{APP_TITLE} — Informe completo", styles),
        Paragraph(f"Módulos: {', '.join(MODULE_BUILDERS[m][0] for m in modules if m in MODULE_BUILDERS)}", styles["Normal"]),
        Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]),
        PageBreak(),
    ]
    for key in modules:
        if key not in MODULE_BUILDERS:
            continue
        label, xlsx_fn, pdf_fn = MODULE_BUILDERS[key]
        story.append(_pdf_section_title(label, styles))
        story.append(Spacer(1, 0.2 * cm))
        if key == "dashboard":
            stats = get_dashboard_stats(db)
            story.append(_pdf_table([
                ["Indicador", "Valor"],
                ["Órdenes activas", stats.active_count],
                ["Kg planificados", stats.total_planned_kg],
                ["Kg producidos", stats.total_produced_kg],
            ], [8 * cm, 5 * cm]))
            if stats.machine_load:
                labels = [r["machine"] for r in stats.machine_load]
                values = [r["working_days"] for r in stats.machine_load]
                story.append(_fig_to_image(_build_chart_figure(labels, values, "Carga máquinas")))
        elif key in ("active_orders", "completed", "gantt"):
            status = None if key == "gantt" else (ItemStatus.ACTIVA.value if key == "active_orders" else ItemStatus.TERMINADA.value)
            if key == "gantt":
                status = ItemStatus.ACTIVA.value
            items = _fetch_orders(db, status)[:80]
            story.append(_pdf_table([
                ["Orden", "Cliente", "Máquina", "Inicio", "Fin", "Kg"],
                *[[i.get("order_number"), i.get("customer"), i.get("machine_name"), i.get("start_date"), i.get("finish_date"), i.get("kg_totales")] for i in items],
            ]))
        elif key == "optimize":
            preview = build_optimization_preview(db)
            cur, opt = preview["current"], preview["optimized"]
            story.append(_pdf_table([
                ["Métrica", "Actual", "Optimizado"],
                ["Tardías", cur["late"], opt["late"]],
                ["Cambios", cur["total_changeovers"], opt["total_changeovers"]],
            ]))
        elif key == "materials":
            mats = db.query(MaterialConfig).order_by(MaterialConfig.material).all()
            story.append(_pdf_table([["Material", "Shrinking"]] + [[m.material, m.shrinking] for m in mats]))
        elif key == "machines":
            machs = db.query(MachineConfig).order_by(MachineConfig.name).all()
            story.append(_pdf_table([["Máquina", "m/turno", "Activa"]] + [[m.name, m.mts_per_shift, "Sí" if m.active else "No"] for m in machs]))
        story.append(PageBreak())
    doc.build(story)
    return buf.getvalue()


def build_complete_zip(db: Session, modules: Sequence[str], lang: str = "es") -> bytes:
    buf = BytesIO()
    stamp = _today_stamp()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        xlsx = build_complete_workbook(db, modules, lang=lang)
        zf.writestr(f"gantt_produccion_completo_{stamp}.xlsx", workbook_to_bytes(xlsx))
        zf.writestr(f"gantt_produccion_completo_{stamp}.pdf", build_complete_pdf(db, modules, lang=lang))
        for key in modules:
            if key not in MODULE_BUILDERS:
                continue
            label, xlsx_fn, pdf_fn = MODULE_BUILDERS[key]
            safe = label.lower().replace(" ", "_")
            zf.writestr(f"{safe}_{stamp}.xlsx", workbook_to_bytes(xlsx_fn(db, lang=lang)))
            zf.writestr(f"{safe}_{stamp}.pdf", pdf_fn(db, lang=lang))
    return buf.getvalue()
